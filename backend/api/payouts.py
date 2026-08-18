from fastapi import APIRouter, Query, HTTPException, Response
from typing import Optional, List, Dict, Any
import sqlite3
import pandas as pd
import math
import os
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.data_processor import LOCAL_DB_PATH, load_data, load_payouts_data

router = APIRouter(prefix="/api/payouts", tags=["Payouts Reconciliation"])


_CLASSIFIED_PAYOUTS_CACHE = None
_CLASSIFICATION_MATRIX_CACHE = None

def invalidate_payouts_cache():
    """Invalidates the in-memory cache for payout reconciliation."""
    global _CLASSIFIED_PAYOUTS_CACHE, _CLASSIFICATION_MATRIX_CACHE
    _CLASSIFIED_PAYOUTS_CACHE = None
    _CLASSIFICATION_MATRIX_CACHE = None

def clean_mojibake_text(text: Any) -> str:
    """Repairs common UTF-8 mojibake, curly quotes, and unwanted encoding artifacts."""
    if not isinstance(text, str) or pd.isna(text):
        return "" if pd.isna(text) else str(text)
    
    s = str(text)
    replacements = {
        'â€™': "'",
        'â€˜': "'",
        'â€œ': '"',
        'â€\x9d': '"',
        'â€': '"',
        'â€“': '-',
        'â€”': '-',
        'Â': '',
        '\xa0': ' ',
        '\xad': '',
        '\u200b': '',
        '\ufeff': '',
        '\ufffd': '',
        '\x81': 'a',
        '’': "'",
        '‘': "'",
        '“': '"',
        '”': '"',
        '–': '-',
        '—': '-',
        'ā': 'a',
        'ū': 'u',
        'ī': 'i',
        'Abū': 'Abu'
    }
    for k, v in replacements.items():
        if k in s:
            s = s.replace(k, v)
            
    if any(c in s for c in ['â', 'Ã']):
        try:
            s = s.encode('latin1').decode('utf-8')
        except Exception:
            pass
            
    for k, v in replacements.items():
        if k in s:
            s = s.replace(k, v)
            
    if 'Their Right Upon Us' in s:
        s = 'Their Right Upon Us | حقهن علينا'
            
    return re.sub(r'[ \t]+', ' ', s).strip()

def _get_classification_matrix_dict() -> Dict[str, Dict[str, str]]:
    global _CLASSIFICATION_MATRIX_CACHE
    if _CLASSIFICATION_MATRIX_CACHE is not None:
        return _CLASSIFICATION_MATRIX_CACHE
    try:
        conn = sqlite3.connect(LOCAL_DB_PATH, timeout=5.0)
        matrix_df = pd.read_sql_query("SELECT campaign_name, heading, sub_heading, country, code, zakat_eligibility FROM campaign_classifications", conn)
        conn.close()
        if not matrix_df.empty:
            for c in ["campaign_name", "heading", "sub_heading", "country", "code", "zakat_eligibility"]:
                if c in matrix_df.columns:
                    matrix_df[c] = matrix_df[c].apply(clean_mojibake_text)
            _CLASSIFICATION_MATRIX_CACHE = {clean_mojibake_text(c).strip().lower(): r for c, r in zip(matrix_df["campaign_name"], matrix_df.to_dict('records'))}
            return _CLASSIFICATION_MATRIX_CACHE
    except Exception as e:
        print(f"[Matrix Overlay Notice]: {e}")
    return {}

def _get_payout_data_from_db(force_reload: bool = False):
    global _CLASSIFIED_PAYOUTS_CACHE
    if not force_reload and _CLASSIFIED_PAYOUTS_CACHE is not None and not _CLASSIFIED_PAYOUTS_CACHE.empty:
        return _CLASSIFIED_PAYOUTS_CACHE

    try:
        df_p = load_payouts_data(force_reload=force_reload)
        if df_p is not None and not df_p.empty:
            df = df_p.copy()
            # Normalize column aliases and clean text artifacts
            c_name = df.get("Campaign Name", pd.Series("Unassigned Campaign", index=df.index)).fillna("Unassigned Campaign").apply(clean_mojibake_text)
            df["campaign_name"] = c_name
            df["Campaign Name"] = c_name
            df["row_type"] = df.get("Type", df.get("Transaction Type", pd.Series("donation", index=df.index))).fillna("donation").astype(str).str.lower()
            df["gross_amt"] = pd.to_numeric(df.get("Total Online Donation Gross Amount in Settled Currency", 0.0), errors="coerce").fillna(0.0)
            df["fee_amt"] = pd.to_numeric(df.get("Total Processing Fees Paid by CC In Settled Currency", 0.0), errors="coerce").fillna(0.0)
            df["net_amt"] = pd.to_numeric(df.get("Total Online Donations Net Amount in Settled Currency", 0.0), errors="coerce").fillna(0.0)
            
            t_ids = df.get("Transfer ID", pd.Series("N/A", index=df.index)).fillna("N/A").astype(str).str.replace(".0", "", regex=False).str.strip()
            df["transfer_id"] = t_ids
            df["Transfer ID"] = t_ids

            # Filter out invalid dummy rows or missing Transfer IDs
            valid_payout_mask = ~df["transfer_id"].str.lower().isin(["n/a", "nan", "none", ""]) & (df["campaign_name"] != "Unassigned Campaign")
            df = df[valid_payout_mask].copy()

            df["heading"] = df.get("Heading", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").apply(clean_mojibake_text)
            df["sub_heading"] = df.get("Sub-Heading", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").apply(clean_mojibake_text)
            df["country"] = df.get("Country", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").apply(clean_mojibake_text)
            df["code"] = df.get("Code", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").apply(clean_mojibake_text)
            df["zakat"] = df.get("Zakat Eligibility", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").apply(clean_mojibake_text)
            df["settlement_currency"] = df.get("Settlement Currency", pd.Series("GBP", index=df.index)).fillna("GBP").astype(str).str.strip().str.upper()
            df["Settlement Currency"] = df["settlement_currency"]

            # Overlay matrix classifications
            rule_dict = _get_classification_matrix_dict()
            if rule_dict:
                c_keys = df["campaign_name"].astype(str).str.strip().str.lower()
                for f, db_f in [("heading", "heading"), ("sub_heading", "sub_heading"), ("country", "country"), ("code", "code"), ("zakat", "zakat_eligibility")]:
                    col_map = {k: clean_mojibake_text(v[db_f]) for k, v in rule_dict.items() if db_f in v and str(v[db_f]).lower() not in ["", "nan", "none", "unassigned"]}
                    mapped = c_keys.map(col_map)
                    valid_m = mapped.notna()
                    if valid_m.any():
                        df.loc[valid_m, f] = mapped[valid_m]

            _CLASSIFIED_PAYOUTS_CACHE = df
            return _CLASSIFIED_PAYOUTS_CACHE

        # Fallback to donations cache if payout_settlements table is empty
        df_don = load_data()
        if not df_don.empty:
            p_mask = df_don.get("Platform", pd.Series("", index=df_don.index)).astype(str).str.lower().str.contains("payout") | df_don.get("Transfer ID", pd.Series(None, index=df_don.index)).notna()
            df_sub = df_don[p_mask]
            if not df_sub.empty:
                df = df_sub.copy()
                c_name = df.get("Campaign Name", pd.Series("Unassigned Campaign", index=df.index)).fillna("Unassigned Campaign")
                df["campaign_name"] = c_name
                df["Campaign Name"] = c_name
                df["row_type"] = df.get("Type", df.get("Transaction Type", pd.Series("donation", index=df.index))).fillna("donation").astype(str).str.lower()
                df["gross_amt"] = pd.to_numeric(df.get("Total Online Donation Gross Amount in Settled Currency", 0.0), errors="coerce").fillna(0.0)
                df["fee_amt"] = pd.to_numeric(df.get("Total Processing Fees Paid by CC In Settled Currency", 0.0), errors="coerce").fillna(0.0)
                df["net_amt"] = pd.to_numeric(df.get("Total Online Donations Net Amount in Settled Currency", 0.0), errors="coerce").fillna(0.0)
                
                t_ids = df.get("Transfer ID", pd.Series("N/A", index=df.index)).fillna("N/A").astype(str).str.replace(".0", "", regex=False).str.strip()
                df["transfer_id"] = t_ids
                df["Transfer ID"] = t_ids

                valid_payout_mask = ~df["transfer_id"].str.lower().isin(["n/a", "nan", "none", ""]) & (df["campaign_name"] != "Unassigned Campaign")
                df = df[valid_payout_mask].copy()

                df["heading"] = df.get("Heading", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").astype(str)
                df["sub_heading"] = df.get("Sub-Heading", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").astype(str)
                df["country"] = df.get("Country", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").astype(str)
                df["code"] = df.get("Code", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").astype(str)
                df["zakat"] = df.get("Zakat Eligibility", pd.Series("Unassigned", index=df.index)).fillna("Unassigned").astype(str)
                df["settlement_currency"] = df.get("Settlement Currency", pd.Series("GBP", index=df.index)).fillna("GBP").astype(str).str.strip().str.upper()
                df["Settlement Currency"] = df["settlement_currency"]
                _CLASSIFIED_PAYOUTS_CACHE = df
                return _CLASSIFIED_PAYOUTS_CACHE

        return pd.DataFrame()
    except Exception as e:
        print(f"[Error] Reading cached payout data: {e}")
        return pd.DataFrame()


def _generate_disbursement_summary(df_curr: pd.DataFrame, currency_name: str) -> Dict[str, Any]:
    """Generates exact Disbursement Summary matching the Finance Team ledger format."""
    donations = df_curr[df_curr["row_type"] == "donation"]
    refunds = df_curr[df_curr["row_type"] == "refund"]
    adjustments = df_curr[df_curr["row_type"] == "adjustment"]
    reserves = df_curr[df_curr["row_type"] == "reserve"]
    fx = df_curr[df_curr["row_type"] == "fx"]
    payouts = df_curr[df_curr["row_type"] == "payout"]

    gross_donations = float(donations["gross_amt"].sum())
    gross_donations_count = int(len(donations))

    refunds_amt = float(refunds["gross_amt"].sum())
    refunds_count = int(len(refunds))

    net_sales = float(gross_donations + refunds_amt)
    net_sales_count = int(gross_donations_count - refunds_count)

    processing_fees = float(df_curr["fee_amt"].sum())
    non_processing_fees = 0.0

    manual_adjustments = float(adjustments["gross_amt"].sum())
    manual_adjustments_count = int(len(adjustments))

    reserve_adjustment = float(reserves["gross_amt"].sum())
    reserve_adjustment_count = int(len(reserves))

    fx_amt = float(fx["gross_amt"].sum())
    fx_count = int(len(fx))

    if not payouts.empty:
        total_disbursement = float(abs(payouts["net_amt"].sum()))
        total_disbursement_count = int(len(payouts))
    else:
        total_disbursement = float(net_sales - processing_fees + manual_adjustments + reserve_adjustment + fx_amt)
        total_disbursement_count = int(len(payouts))

    return {
        "currency": currency_name,
        "gross_donations": round(gross_donations, 2),
        "gross_donations_count": gross_donations_count,
        "refunds": round(refunds_amt, 2),
        "refunds_count": refunds_count,
        "refunds_failed": 0.0,
        "refunds_failed_count": 0,
        "chargebacks": 0.0,
        "chargebacks_count": 0,
        "chargebacks_reversed": 0.0,
        "chargebacks_reversed_count": 0,
        "net_sales": round(net_sales, 2),
        "net_sales_count": net_sales_count,
        "processing_fees": round(-abs(processing_fees) if processing_fees != 0 else 0.0, 2),
        "non_processing_fees": 0.0,
        "manual_adjustments": round(manual_adjustments, 2),
        "manual_adjustments_count": manual_adjustments_count,
        "reserve_adjustment": round(reserve_adjustment, 2),
        "reserve_adjustment_count": reserve_adjustment_count,
        "foreign_exchange": round(fx_amt, 2),
        "foreign_exchange_count": fx_count,
        "total_disbursement": round(total_disbursement, 2),
        "total_disbursement_count": total_disbursement_count
    }


def _generate_ledger_breakdown(df: pd.DataFrame) -> List[Dict[str, Any]]:
    DESCRIPTIONS = {
        "donation": "Gross donor contributions received",
        "payout": "Bank transfer batches paid out to charity account",
        "reserve": "Rolling reserve hold funds",
        "fx": "Foreign exchange conversion adjustments",
        "adjustment": "Settlement / manual account adjustment",
        "refund": "Donor transaction refund deductions"
    }
    TYPE_ORDER = ["donation", "payout", "reserve", "fx", "adjustment", "refund"]

    if df is None or df.empty or "row_type" not in df.columns:
        return []

    breakdown = []
    found_types = set(df["row_type"].unique())
    ordered_types = [t for t in TYPE_ORDER if t in found_types] + [t for t in found_types if t not in TYPE_ORDER]

    for r_type in ordered_types:
        sub = df[df["row_type"] == r_type]
        g_amt = float(sub["gross_amt"].sum())
        f_amt = float(sub["fee_amt"].sum())
        n_amt = float(sub["net_amt"].sum())
        breakdown.append({
            "row_type": r_type.capitalize(),
            "count": int(len(sub)),
            "gross_amount": round(g_amt, 2),
            "processing_fees": round(f_amt, 2),
            "net_amount": round(n_amt, 2),
            "description": DESCRIPTIONS.get(r_type, f"Transaction records for {r_type}")
        })

    return breakdown


@router.get("/summary")
def get_payouts_summary(
    currency: Optional[str] = Query("GBP", description="Filter by settlement currency: GBP, USD, or ALL"),
    batch: Optional[str] = Query("ALL", description="Filter by specific Transfer ID / Batch"),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    campaign_search: Optional[str] = None,
    transfer_id: Optional[str] = None
):
    """Returns top KPI metrics, Finance Disbursement Summary, and Accounting Ledger breakdown."""
    try:
        df_all = _get_payout_data_from_db()
        if df_all.empty:
            return {
                "currency": currency or "GBP",
                "batch": batch or "ALL",
                "total_gross": 0.0,
                "total_fees": 0.0,
                "total_reserves": 0.0,
                "net_payout": 0.0,
                "total_transactions": 0,
                "settled_donations_count": 0,
                "disbursement_summary": {},
                "ledger_breakdown": [],
                "available_currencies": ["GBP", "USD"]
            }

        df = df_all.copy()

        # Currency Filter
        curr_selected = (currency or "GBP").strip().upper()
        if curr_selected in ["GBP", "USD"]:
            df = df[df["settlement_currency"] == curr_selected]
        else:
            curr_selected = "ALL"

        # Batch / Transfer ID Filter
        target_batch = str(transfer_id or batch or "ALL").strip()
        if target_batch and target_batch.upper() != "ALL":
            target_clean = target_batch.replace(".0", "").replace("#", "").strip().lower()
            df = df[df["transfer_id"].astype(str).str.replace(".0", "").str.strip().str.lower() == target_clean]

        # Date Filters
        if start_date and str(start_date).strip():
            df = df[pd.to_datetime(df["Created Date (UTC)"], errors="coerce") >= pd.to_datetime(start_date)]
        if end_date and str(end_date).strip():
            df = df[pd.to_datetime(df["Created Date (UTC)"], errors="coerce") <= pd.to_datetime(end_date)]

        donations_df = df[df["row_type"] == "donation"]
        payouts_df = df[df["row_type"] == "payout"]
        reserves_df = df[df["row_type"] == "reserve"]

        total_gross = float(donations_df["gross_amt"].sum())
        total_fees = float(df["fee_amt"].sum())
        total_reserves = float(abs(reserves_df["net_amt"].sum()))
        
        net_payout = float(abs(payouts_df["net_amt"].sum())) if not payouts_df.empty else float(donations_df["net_amt"].sum() - total_fees)
        
        disbursement_summary = _generate_disbursement_summary(df, curr_selected)
        ledger_breakdown = _generate_ledger_breakdown(df)

        return {
            "currency": curr_selected,
            "batch": target_batch,
            "total_gross": round(total_gross, 2),
            "total_fees": round(total_fees, 2),
            "total_reserves": round(total_reserves, 2),
            "net_payout": round(net_payout, 2),
            "total_transactions": len(df),
            "settled_donations_count": len(donations_df),
            "disbursement_summary": disbursement_summary,
            "ledger_breakdown": ledger_breakdown,
            "available_currencies": ["GBP", "USD", "ALL"]
        }

    except Exception as e:
        print(f"[Error] Payout summary error: {e}")
        return {
            "currency": currency or "GBP",
            "batch": batch or "ALL",
            "total_gross": 0.0,
            "total_fees": 0.0,
            "total_reserves": 0.0,
            "net_payout": 0.0,
            "total_transactions": 0,
            "settled_donations_count": 0,
            "disbursement_summary": {},
            "ledger_breakdown": [],
            "available_currencies": ["GBP", "USD"]
        }


@router.get("/ledger-breakdown")
def get_payout_ledger_breakdown(
    currency: Optional[str] = Query("GBP", description="Filter by settlement currency: GBP, USD, or ALL"),
    batch: Optional[str] = Query("ALL", description="Filter by specific Transfer ID / Batch")
):
    """Returns complete row-type accounting ledger audit table."""
    try:
        df = _get_payout_data_from_db()
        curr_selected = (currency or "GBP").strip().upper()
        if curr_selected in ["GBP", "USD"]:
            df = df[df["settlement_currency"] == curr_selected]

        target_batch = str(batch or "ALL").strip()
        if target_batch and target_batch.upper() != "ALL":
            target_clean = target_batch.replace(".0", "").replace("#", "").strip().lower()
            df = df[df["transfer_id"].astype(str).str.replace(".0", "").str.strip().str.lower() == target_clean]

        return {
            "currency": curr_selected,
            "batch": target_batch,
            "ledger": _generate_ledger_breakdown(df),
            "disbursement_summary": _generate_disbursement_summary(df, curr_selected)
        }
    except Exception as e:
        print(f"[Error] Ledger breakdown error: {e}")
        return {"currency": currency or "GBP", "batch": batch or "ALL", "ledger": [], "disbursement_summary": {}}


@router.get("/batches")
def get_payout_batches(
    currency: Optional[str] = Query("GBP", description="Filter by settlement currency"),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=250),
    search: Optional[str] = ""
):
    """Returns paginated list of payout settlement batches grouped by Transfer ID."""
    try:
        df = _get_payout_data_from_db()
        p_val = page if isinstance(page, int) else 1
        ps_val = page_size if isinstance(page_size, int) else 25
        curr_selected = str(currency).strip().upper() if isinstance(currency, str) else "GBP"
        search_val = str(search).strip().lower() if isinstance(search, str) else ""

        if df.empty:
            return {"total_batches": 0, "page": p_val, "page_size": ps_val, "batches": [], "currency": curr_selected}

        if curr_selected in ["GBP", "USD"]:
            df = df[df["settlement_currency"] == curr_selected]

        # Filter out empty or N/A transfer IDs
        df = df[~df["transfer_id"].astype(str).str.lower().isin(["n/a", "nan", "none", ""])].copy()

        if df.empty:
            return {"total_batches": 0, "page": p_val, "page_size": ps_val, "batches": [], "currency": curr_selected}

        if search_val:
            mask = df["transfer_id"].astype(str).str.lower().str.contains(search_val, na=False) | df["campaign_name"].astype(str).str.lower().str.contains(search_val, na=False)
            df = df[mask]

        grouped = df.groupby("transfer_id").apply(lambda g: pd.Series({
            "created_date": str(g["Created Date (UTC)"].dropna().min() or "N/A"),
            "gross_amount": round(float(g[g["row_type"] == "donation"]["gross_amt"].sum()), 2),
            "processing_fees": round(float(g["fee_amt"].sum()), 2),
            "transfer_amount": round(float(abs(g[g["row_type"] == "payout"]["net_amt"].sum()) or g[g["row_type"] == "donation"]["net_amt"].sum()), 2),
            "campaigns_count": int(g["campaign_name"].nunique()),
            "donations_count": int((g["row_type"] == "donation").sum()),
            "currency": str(g["settlement_currency"].iloc[0] if "settlement_currency" in g.columns else "GBP")
        })).reset_index()

        # Sort batches by created_date descending or transfer_id descending
        grouped = grouped.sort_values(by=["created_date", "transfer_id"], ascending=[False, False])

        total_batches = len(grouped)
        total_pages = max(1, math.ceil(total_batches / ps_val))
        p_val = min(p_val, total_pages)
        start_idx = (p_val - 1) * ps_val
        end_idx = min(start_idx + ps_val, total_batches)

        page_batches = grouped.iloc[start_idx:end_idx].to_dict(orient="records")

        return {
            "total_batches": total_batches,
            "page": p_val,
            "page_size": ps_val,
            "total_pages": total_pages,
            "batches": page_batches,
            "currency": curr_selected
        }

    except Exception as e:
        print(f"[Error] Payout batches error: {e}")
        return {"total_batches": 0, "page": 1, "page_size": 25, "batches": [], "currency": "GBP"}


@router.get("/campaign-breakdown")
def get_campaign_payout_breakdown(
    currency: Optional[str] = Query("GBP", description="Filter by settlement currency"),
    batch: Optional[str] = Query("ALL", description="Filter by specific Transfer ID / Batch"),
    search: Optional[str] = ""
):
    """Returns classification code-level hierarchical breakdown with nested campaigns and individual campaign metrics (Optimized Vectorized Aggregation)."""
    try:
        df = _get_payout_data_from_db()
        if df.empty:
            return {"code_groups": [], "campaigns": [], "total_codes": 0, "total_campaigns": 0, "currency": currency or "GBP", "batch": batch or "ALL"}

        curr_selected = str(currency).strip().upper() if isinstance(currency, str) else "GBP"
        if curr_selected in ["GBP", "USD"]:
            df = df[df["settlement_currency"] == curr_selected]

        target_batch = str(batch or "ALL").strip()
        if target_batch and target_batch.upper() != "ALL":
            target_clean = target_batch.replace(".0", "").replace("#", "").strip().lower()
            df = df[df["transfer_id"].astype(str).str.replace(".0", "").str.strip().str.lower() == target_clean]

        search_val = str(search).strip().lower() if isinstance(search, str) else ""

        # 1. Vectorized Aggregation per Campaign & Row Type
        camp_agg = df.groupby(["campaign_name", "row_type"]).agg(
            gross=("gross_amt", "sum"),
            fee=("fee_amt", "sum"),
            net=("net_amt", "sum"),
            count=("gross_amt", "count")
        ).reset_index()

        meta = df.groupby("campaign_name").first()[["heading", "sub_heading", "country", "code", "zakat"]].to_dict('index')

        campaign_records = []
        for cname, g in camp_agg.groupby("campaign_name"):
            c_meta = meta.get(cname, {})
            row_dict = {r["row_type"]: r for _, r in g.iterrows()}
            
            don = row_dict.get("donation", {})
            ref = row_dict.get("refund", {})
            
            don_gross = float(don.get("gross", 0.0))
            ref_gross = float(ref.get("gross", 0.0))
            g_amt = round(don_gross + ref_gross, 2)
            
            f_amt = round(float(g["fee"].sum()), 2)
            
            don_net = float(don.get("net", 0.0))
            ref_net = float(ref.get("net", 0.0))
            fx_net = float(row_dict.get("fx", {}).get("gross", 0.0))
            adj_net = float(row_dict.get("adjustment", {}).get("gross", 0.0))
            res_net = float(row_dict.get("reserve", {}).get("gross", 0.0))
            
            if curr_selected == "USD":
                t_amt = round(don_net + ref_net, 2)
            else:
                # Comprehensive Net Settlement = Donations Net + Refunds Net + Foreign FX Inflow + Manual Adjustments + Reserve Holds/Releases
                t_amt = round(don_net + ref_net + fx_net + adj_net + res_net, 2)

            fee_pct = round((f_amt / g_amt * 100.0), 2) if g_amt > 0 else 0.0
            
            don_cnt = int(don.get("count", 0))
            ref_cnt = int(ref.get("count", 0))

            campaign_records.append({
                "campaign_name": str(cname),
                "gross_amount": g_amt,
                "processing_fees": f_amt,
                "transfer_amount": t_amt,
                "fee_percentage": fee_pct,
                "heading": str(c_meta.get("heading", "Unassigned")),
                "sub_heading": str(c_meta.get("sub_heading", "Unassigned")),
                "country": str(c_meta.get("country", "Unassigned")),
                "code": str(c_meta.get("code", "Unassigned")),
                "zakat": str(c_meta.get("zakat", "Unassigned")),
                "donations_count": don_cnt if don_cnt > 0 else (ref_cnt if ref_cnt > 0 else 0)
            })

        campaign_records.sort(key=lambda x: x["gross_amount"], reverse=True)

        # 2. Build Code Groups from campaign_records
        code_map = {}
        for c in campaign_records:
            cd = c["code"]
            if cd not in code_map:
                code_map[cd] = {
                    "code": cd,
                    "heading": c["heading"],
                    "sub_heading": c["sub_heading"],
                    "country": c["country"],
                    "zakat": c["zakat"],
                    "gross_amount": 0.0,
                    "processing_fees": 0.0,
                    "transfer_amount": 0.0,
                    "donations_count": 0,
                    "campaigns": []
                }
            code_map[cd]["gross_amount"] = round(code_map[cd]["gross_amount"] + c["gross_amount"], 2)
            code_map[cd]["processing_fees"] = round(code_map[cd]["processing_fees"] + c["processing_fees"], 2)
            code_map[cd]["transfer_amount"] = round(code_map[cd]["transfer_amount"] + c["transfer_amount"], 2)
            code_map[cd]["donations_count"] += c["donations_count"]
            code_map[cd]["campaigns"].append({
                "campaign_name": c["campaign_name"],
                "gross_amount": c["gross_amount"],
                "processing_fees": c["processing_fees"],
                "transfer_amount": c["transfer_amount"],
                "fee_percentage": c["fee_percentage"],
                "donations_count": c["donations_count"]
            })

        code_groups = list(code_map.values())
        for cg in code_groups:
            cg["campaigns_count"] = len(cg["campaigns"])
            cg["fee_percentage"] = round((cg["processing_fees"] / cg["gross_amount"] * 100.0), 2) if cg["gross_amount"] > 0 else 0.0

        code_groups.sort(key=lambda x: x["gross_amount"], reverse=True)

        # Apply Search Filter
        if search_val:
            filtered_camps = [
                c for c in campaign_records 
                if search_val in c["campaign_name"].lower() 
                or search_val in c["code"].lower() 
                or search_val in c["heading"].lower() 
                or search_val in c["sub_heading"].lower()
                or search_val in c["country"].lower()
            ]

            filtered_codes = []
            for cg in code_groups:
                matching_subs = [
                    sc for sc in cg["campaigns"]
                    if search_val in sc["campaign_name"].lower()
                    or search_val in cg["code"].lower()
                    or search_val in cg["heading"].lower()
                    or search_val in cg["sub_heading"].lower()
                    or search_val in cg["country"].lower()
                ]
                if matching_subs:
                    cg_copy = dict(cg)
                    cg_copy["campaigns"] = matching_subs
                    cg_copy["campaigns_count"] = len(matching_subs)
                    filtered_codes.append(cg_copy)

            return {
                "code_groups": filtered_codes,
                "campaigns": filtered_camps,
                "total_codes": len(filtered_codes),
                "total_campaigns": len(filtered_camps),
                "currency": curr_selected,
                "batch": target_batch
            }

        return {
            "code_groups": code_groups,
            "campaigns": campaign_records,
            "total_codes": len(code_groups),
            "total_campaigns": len(campaign_records),
            "currency": curr_selected,
            "batch": target_batch
        }

    except Exception as e:
        print(f"[Error] Campaign payout breakdown error: {e}")
        return {"code_groups": [], "campaigns": [], "total_codes": 0, "total_campaigns": 0, "currency": currency or "GBP", "batch": batch or "ALL"}


@router.get("/export")
def export_payouts_report(
    currency: Optional[str] = Query("ALL", description="Currency filter"),
    batch: Optional[str] = Query("ALL", description="Filter by Transfer ID / Batch")
):
    """Generates and exports a comprehensive multi-sheet Excel (.xlsx) reconciliation report."""
    try:
        curr_selected = str(currency).strip().upper() if isinstance(currency, str) else "ALL"
        target_batch = str(batch or "ALL").strip()
        
        # 1. Fetch Data
        summary_data = get_payouts_summary(currency=curr_selected, batch=target_batch)
        camp_data = get_campaign_payout_breakdown(currency=curr_selected, batch=target_batch)
        batch_data = get_payout_batches(currency=curr_selected, page_size=1000)
        disb = summary_data.get("disbursement_summary", {})
        ledger_breakdown = summary_data.get("ledger_breakdown", [])

        # 2. Build Workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet

        # Styles
        header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=11)
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB")
        )

        # -----------------------------------------------------------------
        # Sheet 1: Disbursement Summary
        # -----------------------------------------------------------------
        ws1 = wb.create_sheet(title="Disbursement Summary")
        ws1.append(["Disbursement Summary Line Item", "Transaction Count", f"Amount Value ({curr_selected})", "Notes"])
        disb_rows = [
            ["Gross Donations / Contributions", disb.get("gross_donations_count", 0), disb.get("gross_donations", 0.0), "Gross donor contributions received"],
            ["Refunds", disb.get("refunds_count", 0), disb.get("refunds", 0.0), "Donor transaction refund deductions"],
            ["Refunds Failed", disb.get("refunds_failed_count", 0), disb.get("refunds_failed", 0.0), "Failed refund attempts"],
            ["Chargebacks", disb.get("chargebacks_count", 0), disb.get("chargebacks", 0.0), "Disputed donor transactions"],
            ["Chargebacks Reversed", disb.get("chargebacks_reversed_count", 0), disb.get("chargebacks_reversed", 0.0), "Resolved / won chargebacks"],
            ["Net Sales", disb.get("net_sales_count", 0), disb.get("net_sales", 0.0), "Gross donations minus refunds"],
            ["Processing Fees", "", disb.get("processing_fees", 0.0), "Credit card and platform processing fees"],
            ["Non Processing Fees", "", disb.get("non_processing_fees", 0.0), "Other non-processing fees"],
            ["Manual Adjustments Total (Zakat donation fees)", disb.get("manual_adjustments_count", 0), disb.get("manual_adjustments", 0.0), "Manual account adjustments and fees"],
            ["Reserve Adjustment", disb.get("reserve_adjustment_count", 0), disb.get("reserve_adjustment", 0.0), "Rolling reserve funds hold and release"],
            ["Foreign Exchange", disb.get("foreign_exchange_count", 0), disb.get("foreign_exchange", 0.0), "Foreign currency conversions into settlement account"],
            ["Total Disbursement", disb.get("total_disbursement_count", 0), disb.get("total_disbursement", 0.0), "Net amount disbursed to charity bank account"]
        ]
        for r in disb_rows:
            ws1.append(r)

        # -----------------------------------------------------------------
        # Sheet 2: Code Breakdown
        # -----------------------------------------------------------------
        ws2 = wb.create_sheet(title="Code Breakdown")
        ws2.append(["Classification Code", "Heading", "Sub-Heading", "Country", "Zakat Eligibility", "Contributing Campaigns", "Donations Count", f"Gross Raised ({curr_selected})", f"Processing Fees ({curr_selected})", "Fee Ratio (%)", f"Net Settlement ({curr_selected})"])
        for cg in camp_data.get("code_groups", []):
            ws2.append([
                cg.get("code", ""),
                cg.get("heading", ""),
                cg.get("sub_heading", ""),
                cg.get("country", ""),
                cg.get("zakat", ""),
                cg.get("campaigns_count", 0),
                cg.get("donations_count", 0),
                cg.get("gross_amount", 0.0),
                cg.get("processing_fees", 0.0),
                f"{cg.get('fee_percentage', 0.0)}%",
                cg.get("transfer_amount", 0.0)
            ])

        # -----------------------------------------------------------------
        # Sheet 3: Campaign Breakdown
        # -----------------------------------------------------------------
        ws3 = wb.create_sheet(title="Campaign Breakdown")
        ws3.append(["Campaign Name", "Classification Code", "Heading", "Sub-Heading", "Country", "Zakat Eligibility", "Donations Count", f"Gross Amount ({curr_selected})", f"Processing Fees ({curr_selected})", "Fee Ratio (%)", f"Net Settlement ({curr_selected})"])
        for c in camp_data.get("campaigns", []):
            ws3.append([
                c.get("campaign_name", ""),
                c.get("code", ""),
                c.get("heading", ""),
                c.get("sub_heading", ""),
                c.get("country", ""),
                c.get("zakat", ""),
                c.get("donations_count", 0),
                c.get("gross_amount", 0.0),
                c.get("processing_fees", 0.0),
                f"{c.get('fee_percentage', 0.0)}%",
                c.get("transfer_amount", 0.0)
            ])

        # -----------------------------------------------------------------
        # Sheet 4: Transfer Batches
        # -----------------------------------------------------------------
        ws4 = wb.create_sheet(title="Transfer Batches")
        ws4.append(["Transfer ID", "Settlement Date", "Currency", "Campaigns Count", "Donations Count", "Gross Amount", "Processing Fees", "Net Payout"])
        for b in batch_data.get("batches", []):
            ws4.append([
                f"#{b.get('transfer_id', '')}",
                b.get("created_date", ""),
                b.get("currency", ""),
                b.get("campaigns_count", 0),
                b.get("donations_count", 0),
                b.get("gross_amount", 0.0),
                b.get("processing_fees", 0.0),
                b.get("transfer_amount", 0.0)
            ])

        # -----------------------------------------------------------------
        # Sheet 5: Ledger Audit
        # -----------------------------------------------------------------
        ws5 = wb.create_sheet(title="Ledger Audit")
        ws5.append(["Row Type", "Description", "Row Count", "Gross Amount", "Processing Fees", "Net Amount"])
        for l in ledger_breakdown:
            ws5.append([
                l.get("row_type", ""),
                l.get("description", ""),
                l.get("count", 0),
                l.get("gross_amount", 0.0),
                l.get("processing_fees", 0.0),
                l.get("net_amount", 0.0)
            ])

        # Apply Formatting & Column Sizing to All Sheets
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 26

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = regular_font
                    cell.border = thin_border
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = str(cell.value or "")
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 50)

        # 3. Stream Response
        output = io.BytesIO()
        wb.save(output)
        content_bytes = output.getvalue()

        filename = f"payout_reconciliation_report_{curr_selected.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            content=content_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"[Error] Export payouts report error: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel export: {str(e)}")


